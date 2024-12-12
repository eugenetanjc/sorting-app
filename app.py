# Import libraries
import streamlit as st
import hmac
import pandas as pd
import os
import io
from io import BytesIO
import requests
from PIL import Image as PILImage
import base64
import backend
import compile
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl import load_workbook

# Define global variables
image_filename = 'data/ck_logo.png'
with open(image_filename, 'rb') as image_file:
    encoded_image = base64.b64encode(image_file.read()).decode()

# Create dictionaries for input options
year = int(datetime.now().strftime("%Y")) 
week_options = [{'label': str(i), 'value': str(i)} for i in range(1, 53)]
year_options = [{'label': str(i), 'value': str(i)} for i in range(year-1, year+2)]

# Country options
country_options = [{'label': country, 'value': country} for country in [
    'INDIA', 'INDONESIA', 'JAPAN', 'SAUDI ARABIA', 'SINGAPORE', 'SOUTH KOREA', 'THAILAND', 'VIETNAM'
    ]]
countrytype_options = [{'label': countrytype, 'value': countrytype} for countrytype  in ['Hot', 'Cold', 'ANZ']]

def peform_sorting(s_country, s_year, s_week, s_ctype, params_dict):
    global username, input_country, input_ctype

    if not all([s_country, s_year, s_week, s_ctype]):
        st.error("Please fill in all required fields.")

    # Check user input
    print(f'Selected Country: {s_country}, Year: {s_year}, Week: {s_week}, Seasonality: {s_ctype}')

    # Run sorting function
    output_df = backend.sorting(s_country, s_year, s_week, s_ctype, params_dict)
    return output_df

# Function to create a downloadable link for CSV file
def get_csv_download_link(df, filename="sorted_data.xlsx"):
    
    # Loop through the rows of the DataFrame and split the data into sheets based on the 'Category ID' values
    sheets = {}
    for _, row in df.iterrows():
        category_id = row['Category ID']
        if category_id not in sheets:
            sheets[category_id] = pd.DataFrame(columns=df.columns)
        sheets[category_id] = pd.concat([sheets[category_id], row.to_frame().T])

    # Create a new workbook object for the output file
    output_workbook = Workbook()

    # Loop through the sheets and add them to the output workbook
    for category_id, sheet_df in sheets.items():
        sheet_df.drop_duplicates(subset=['Article'], keep='first', inplace=True)
        output_worksheet = output_workbook.create_sheet(category_id)

        # Append column headers
        output_worksheet.append(sheet_df.columns.tolist())
        
        # Append data rows
        for row in sheet_df.itertuples(index=False, name=None):
            output_worksheet.append(row)

    output_workbook.remove(output_workbook['Sheet'])  

    # Save the workbook to a BytesIO object
    excel_buffer = BytesIO()
    output_workbook.save(excel_buffer)
    excel_buffer.seek(0)

    b64 = base64.b64encode(excel_buffer.read()).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">Download Excel file</a>'
    return href
    
def sorting_layout():

    st.markdown(
    f'<div style="text-align: center; margin: 10px;">'
    f'<div><img src="data:image/png;base64,{encoded_image}" style="height: 40%; width: 40%;"/></div>'
    f'<br/>'
    f'<div style="text-align: center; margin-bottom: 15px; display: inline-block; font-family: Bahnschrift; font-weight: bold;">'
    f'<h1>ECOM Sorting App</h1>'
    f'</div>'
    f'</div>',
    unsafe_allow_html=True
    )
    st.markdown('<hr>', unsafe_allow_html=True)

    st.header("Sorting")

    def check_password():
        """Returns `True` if the user had the correct password."""

        def password_entered():
            """Checks whether a password entered by the user is correct."""
            if hmac.compare_digest(st.session_state["password"], st.secrets["password"]):
                st.session_state["password_correct"] = True
                del st.session_state["password"]  # Don't store the password.
            else:
                st.session_state["password_correct"] = False

        # Return True if the password is validated.
        if st.session_state.get("password_correct", False):
            return True

        # Show input for password.
        st.text_input(
            "Password", type="password", on_change=password_entered, key="password"
        )
        if "password_correct" in st.session_state:
            st.error("😕 Password incorrect")
        return False

    if not check_password():
        st.stop()  # Do not continue if check_password is not True.

    # Upload Parameters File
    st.subheader("Upload Parameters File")
    uploaded_file = st.file_uploader("Drag and Drop or Select Parameters file", type=["xlsx", "xls"])
    if uploaded_file is not None:
        xls = pd.ExcelFile(uploaded_file, engine='openpyxl')
        params_dict = {sheet_name: pd.read_excel(xls, sheet_name=sheet_name) for sheet_name in xls.sheet_names}
        st.success("File uploaded successfully!")

    # Select your country
    st.subheader("Select your country")
    country = st.selectbox("Select a country", [opt['value'] for opt in country_options], index=0)

    # Select your year
    st.subheader("Select your year")
    year = st.selectbox("Select a year", [opt['value'] for opt in year_options], index=1)

    # Select your week
    st.subheader("Select your week")
    week = st.selectbox("Select a week", options=[opt['value'] for opt in week_options], index=23)

    # Select your country type
    st.subheader("Select your country type")
    country_type = st.selectbox("Select a country type", options=[opt['value'] for opt in countrytype_options], index=0)

    # Run sorting now button
    if st.button("Run sorting now"):
        st.write("Sorting is being executed...")
        outcome = peform_sorting(country, year, week, country_type, params_dict)
        st.session_state['outcome'] = outcome
        st.markdown(get_csv_download_link(st.session_state['outcome'], f"Output_{country}_W{week}_Y{year}_{country_type}"), unsafe_allow_html=True)
        st.success("Sorting completed, file is ready for download!")

# Define the Streamlit app layout for the Compilation tab
def compile_layout():

    st.markdown(
    f'<div style="text-align: center; margin: 10px;">'
    f'<div><img src="data:image/png;base64,{encoded_image}" style="height: 40%; width: 40%;"/></div>'
    f'<br/>'
    f'<div style="text-align: center; margin-bottom: 15px; display: inline-block; font-family: Bahnschrift; font-weight: bold;">'
    f'<h1>ECOM Sorting App</h1>'
    f'</div>'
    f'</div>',
    unsafe_allow_html=True
    )
    st.markdown('<hr>', unsafe_allow_html=True)

    st.header("Compilation")

    # Input your working file
    st.subheader("Input your working file (should be a .xlsx file) into the area below for compile:")
    uploaded_file = st.file_uploader("Drag and Drop or Select .xlsx working file", type=["xlsx"])
    if uploaded_file:
        st.write("File uploaded!")

    # Compile button
    if st.button("Compile working file"):
        if uploaded_file:
            st.write("Compiling working file...")

            # Load the workbook 
            workbook = openpyxl.load_workbook(uploaded_file)
            dfs = [pd.read_excel(BytesIO(uploaded_file.getvalue()), sheet_name=sheet_name, usecols="A:D", engine='openpyxl') for sheet_name in workbook.sheetnames]
            combined_df = pd.concat(dfs, ignore_index=True)
            
            # Define the maximum number of rows per file
            max_rows_per_file = 2000

            # Group by the category column (replace 'category' with the actual category column name)
            df_grouped = combined_df.groupby('Category ID', sort=False)

            # List to store chunks
            chunks = []
            current_chunk = pd.DataFrame()

            for _, group in df_grouped:
                # If adding this group exceeds the max_rows_per_file, store the current chunk and start a new one
                if len(current_chunk) + len(group) > max_rows_per_file - 1:
                    chunks.append(current_chunk)
                    current_chunk = pd.DataFrame()
                
                # Add the group to the current chunk
                current_chunk = pd.concat([current_chunk, group], ignore_index=True)

            # Add the last chunk if it's not empty
            if not current_chunk.empty:
                chunks.append(current_chunk)

            # Create download links for each chunk
            for i, chunk in enumerate(chunks):
                # Convert the chunk to a CSV string
                csv = chunk.to_csv(index=False)

                # Encode the CSV string as base64
                b64 = base64.b64encode(csv.encode()).decode()

                # Get the current date and format it as a string without dashes
                formatted_date = datetime.now().strftime("%Y%m%d")

                # Create a download link for the chunk
                href = f'<a href="data:file/csv;base64,{b64}" download="categoryposition_{formatted_date}_{i+1:02}.csv">Download CSV file Part {i+1}</a>'

                # Display the download link in Streamlit
                st.markdown(href, unsafe_allow_html=True)
            
            st.success("Compilation complete!")
        else:
            st.error("Please upload at least one file.")

# Main app layout
def serve_layout():
    st.sidebar.title("ECOM Sorting App")
    tab_selection = st.sidebar.radio("Go to", ["Sorting", "Compilation"])

    if tab_selection == "Sorting":
        sorting_layout()

    elif tab_selection == "Compilation":
        compile_layout()

if __name__ == "__main__":
    serve_layout()